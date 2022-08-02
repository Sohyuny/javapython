from re import L
from openpyxl import load_workbook
'''
파일 열때는 load_workbook
파일 생성할때는 workbook
'''

wb=load_workbook("b.xlsx")
ws=wb.active

print("현재 작성된 함수 : ",ws.max_row)
print("현재 작성된 컬럼 수 : ",ws.max_column)

for r in range(1,ws.max_row):
    for c in range(1,ws.max_column):
        print(f"row = {r}, column = {c} = ",ws.cell(row=r,column=c).value,end=" ")
    print()

wb.save("c.xlsx")
wb.close()