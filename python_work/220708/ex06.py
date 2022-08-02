from openpyxl import Workbook

wb=Workbook()
ws=wb.active

ws['A1']='이상한 변호사'
ws['A2']='우영우'
ws.title='고래얘기금지'
ws.sheet_propertie.tapColor="#1072BA"

ms=wb.create_sheet("My sheet")
ms['B1']="똑바로해도 거꾸로해도 우영우"
ms['B2']='토마토 기러기 스위스 인도인 별똥별 우영우'

ms=wb.create_sheet("MY sheet",0)
ms['C1']='...역삼역?'
ms['C2']='그런건 회사에서 하면 안돼'

ms3=wb["MySheet"]

wb.save('a.xlsx')
wb.close()